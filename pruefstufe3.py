#!/usr/bin/env python3
"""
pruefstufe3.py – berechnet Prüfstufe 3 (INAD‑Dichte) für beliebiges Zeitfenster.

Beispielaufruf:
  python pruefstufe3.py \
        --inad  "INAD Tabelle .xlsm" \
        --bazl  "BAZL-Daten.xlsx" \
        --start 2024-07-01 --end 2024-12-31 \
        --inadmin 6 --map partner.csv --oe 0.051
"""

import argparse, pandas as pd
from datetime import datetime
import openpyxl, collections, csv

# ---------------------- Einstellungen ----------------------------------
EXCLUDE_CODES = {'B1n','B2n','C4n','C5n','C8','D1n','D2n','E','F1n','G','H','I'}

# ---------------------- Argument‑Parser --------------------------------
def parse_args():
    ap = argparse.ArgumentParser()
    ap.add_argument('--inad',  required=True, help='INAD Excel (xls/xlsm)')
    ap.add_argument('--bazl',  required=True, help='BAZL Excel (xlsx)')
    ap.add_argument('--start', required=True, help='YYYY-MM-DD')
    ap.add_argument('--end',   required=True, help='YYYY-MM-DD')
    ap.add_argument('--inadmin', type=int, default=6, help='Schwelle ≥ INAD')
    ap.add_argument('--oe',    type=float, default=None, help='feste Ø-Schwelle in ‰')
    ap.add_argument('--map',   default=None, help='CSV PartnerMapping Carrier;LastStop;Partner')
    return ap.parse_args()

def load_partner_map(path):
    mp = collections.defaultdict(list)
    if not path: return mp
    with open(path, newline='') as f:
        for carrier,lstop,partner in csv.reader(f, delimiter=';'):
            mp[(carrier.strip(), lstop.strip())].append(partner.strip())
    return mp

def excel_serial_to_date(n):
    return pd.to_datetime('1899-12-30') + pd.to_timedelta(int(n), 'D')

# ======================================================================
def main():
    args = parse_args()
    start, end = pd.to_datetime(args.start), pd.to_datetime(args.end)
    partner_map = load_partner_map(args.map)

    # ---------- 1) BAZL PAX --------------------------------------------
    cols = ['Airline Code (IATA)', 'Flughafen (IATA)',
            'Passagiere / Passagers', 'Jahr', 'Monat']
    df_b = pd.read_excel(args.bazl, sheet_name='BAZL-Daten', usecols=cols)

    # robust von Seriennummer, Zahl oder Datum → Monats‑Datum
    def month_to_date(row):
        m = row['Monat']
        if isinstance(m, (int, float)) and m > 1000:
            return excel_serial_to_date(m)
        if isinstance(m, (int, float)):
            return datetime(int(row['Jahr']), int(m), 1)
        if isinstance(m, datetime):
            return datetime(m.year, m.month, 1)
        return pd.NaT

    df_b['Date'] = df_b.apply(month_to_date, axis=1)
    df_b = df_b[(df_b['Date'] >= start) & (df_b['Date'] <= end)]

    pax = (df_b.groupby(['Airline Code (IATA)', 'Flughafen (IATA)'])
                ['Passagiere / Passagers'].sum())

    # ---------- 2) INAD Fälle ------------------------------------------
    wb = openpyxl.load_workbook(args.inad, read_only=True, data_only=True)
    ws = wb['INAD-Tabelle']
    hdr = {c.value: i for i, c in enumerate(ws[1])}
    idx_air   = hdr['Fluggesellschaft']
    idx_last  = hdr['Abflugort (last stop)']
    idx_year  = hdr['Jahr']
    idx_month = hdr['Monat']
    idx_det   = 18   # Spalte S

    air_cnt, rt_cnt = collections.Counter(), collections.Counter()

    for row in ws.iter_rows(min_row=2, values_only=True):
        yr  = row[idx_year]
        mon = row[idx_month]
        # leer oder nicht‑numerisch überspringen
        if yr is None or mon is None:
            continue
        try:
            yr = int(yr); mon = int(mon)
        except ValueError:
            continue
        dt = datetime(yr, mon, 1)
        if dt < start or dt > end:
            continue

        code = (row[idx_det] or '').strip()
        if not code or code in EXCLUDE_CODES:
            continue

        air = (row[idx_air]  or '').strip()
        lst = (row[idx_last] or '').strip()
        air_cnt[air] += 1
        rt_cnt[(air, lst)] += 1

    # ---------- 3) Prüfstufe 2 Filter ----------------------------------
    sel_air = {a for a,c in air_cnt.items() if c >= args.inadmin}
    sel_rt  = {k:v for k,v in rt_cnt.items() if k[0] in sel_air and v >= args.inadmin}

    # ---------- 4) Prüfstufe 3 Berechnung ------------------------------
    rows = []
    for (air,lst), inad in sel_rt.items():
        p = pax.get((air,lst), 0)
        for partner in partner_map.get((air,lst), []):
            p += pax.get((partner,lst), 0)
        dens = (inad/p*1000) if p else None
        rows.append([air, lst, inad, int(p), dens])

    res = pd.DataFrame(rows, columns=['Carrier','LastStop','INAD','PAX','Density‰'])

    # ---------- 5) Ø-Wert ----------------------------------------------
    thresh = args.oe if args.oe is not None else res['Density‰'].mean()
    res['Flag'] = res['Density‰'].apply(lambda x: '≥Ø' if x>=thresh else '<Ø')
    res.sort_values(['Flag','INAD'], ascending=[False,False], inplace=True)

    # ---------- 6) Ausgabe ---------------------------------------------
    out = f"Pruefstufe3_{start:%Y%m%d}_{end:%Y%m%d}.xlsx"
    with pd.ExcelWriter(out) as w:
        res.to_excel(w, index=False, sheet_name='Pruefstufe3')
    print(f"✓ Fertig. Ø-Wert = {thresh:.3f} ‰  →  {out}")

# ======================================================================
if __name__ == "__main__":
    main()
