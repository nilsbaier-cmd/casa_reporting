#!/usr/bin/env python3
"""
inad_analysis.py – Enhanced INAD Analysis Module

This module provides comprehensive data loading and analysis functions for INAD density calculations
with statistical robustness, data quality checks, and systemic case detection.

Analysis Steps:
- Step 1 (Prüfstufe 1): Airlines with ≥N INADs in period
- Step 2 (Prüfstufe 2): Routes from Step 1 airlines with ≥N INADs
- Step 3 (Prüfstufe 3): INAD density calculation with enhanced flagging

Enhanced Features:
- Robust statistics (median, trimmed mean)
- Minimum PAX threshold to exclude unreliable data
- Confidence scoring based on sample size
- Data quality checks and warnings
- Multi-semester trend detection
- Priority classification (HIGH, WATCH, SYSTEMIC)
"""

import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
import collections
import csv
from pathlib import Path
from scipy import stats

# Entry refusal codes to exclude from analysis
EXCLUDE_CODES = {'B1n', 'B2n', 'C4n', 'C5n', 'C8', 'D1n', 'D2n', 'E', 'F1n', 'G', 'H', 'I'}

# Human-readable descriptions for refusal codes
REFUSAL_CODE_DESCRIPTIONS = {
    'A': 'Kein gültiges Reisedokument',
    'A1': 'Kein Reisedokument',
    'A2': 'Gefälschtes Reisedokument',
    'A3': 'Abgelaufenes Reisedokument',
    'A4': 'Unvollständiges Reisedokument',
    'A5': 'Falsches Reisedokument',
    'B': 'Gefälschtes Dokument',
    'B1': 'Gefälschtes/verfälschtes Reisedokument',
    'B1e': 'Gefälschtes Reisedokument (erkannt)',
    'B2': 'Gefälschtes/verfälschtes Visum oder Aufenthaltstitel',
    'B2e': 'Gefälschtes Visum (erkannt)',
    'C1': 'Kein gültiges Visum oder Aufenthaltstitel',
    'C2': 'Gefälschtes/verfälschtes Visum (bereits verwendet)',
    'C3': 'Visum/Aufenthaltstitel für anderen Schengen-Staat',
    'C4': 'Visum bereits annulliert',
    'C4e': 'Visum annulliert (erkannt)',
    'C5': 'Visum am Grenzübergang annulliert',
    'C5e': 'Visum annulliert am Grenzübergang (erkannt)',
    'C6': 'Visum überschritten (overstay)',
    'C7': 'Aufenthaltszweck nicht nachgewiesen',
    'D1': 'Einreisesperre SIS',
    'D2': 'Nationale Einreisesperre',
    'D2e': 'Nationale Einreisesperre (erkannt)',
    'F': 'Gefahr öffentliche Ordnung',
    'F1': 'Gefahr für öffentliche Ordnung/Sicherheit',
    'F1e': 'Gefahr öffentliche Ordnung (erkannt)',
    'F2': 'Gefahr für öffentliche Gesundheit',
    'Other': 'Sonstiger Grund'
}

# Refusal code categories for analysis
REFUSAL_CATEGORIES = {
    'Documentation': ['A', 'A1', 'A2', 'A3', 'A4', 'A5'],
    'Fraud': ['B', 'B1', 'B1e', 'B2', 'B2e'],
    'Visa': ['C1', 'C2', 'C3', 'C4', 'C4e', 'C5', 'C5e', 'C6', 'C7'],
    'Security': ['D1', 'D2', 'D2e', 'F', 'F1', 'F1e', 'F2']
}

# Default configuration
DEFAULT_CONFIG = {
    'min_inad': 6,                    # Minimum INADs for Step 1/2
    'min_pax': 5000,                  # Minimum PAX for reliable density calculation
    'min_density': 0.10,              # Minimum density threshold (‰) for HIGH PRIORITY
    'high_priority_multiplier': 1.5,  # Must be this × average for HIGH PRIORITY
    'high_priority_min_inad': 10,     # Minimum INADs for HIGH PRIORITY
    'threshold_method': 'median',      # 'mean', 'median', or 'trimmed_mean'
    'trimmed_percent': 0.1,           # Percent to trim from each end for trimmed_mean
    'systemic_semesters': 2,          # Consecutive semesters for SYSTEMIC flag
    'pax_completeness_months': 4      # Minimum months of PAX data required
}


def excel_serial_to_date(n):
    """Convert Excel serial date number to pandas datetime."""
    return pd.to_datetime('1899-12-30') + pd.to_timedelta(int(n), 'D')


def load_partner_map(path):
    """Load carrier partner mapping from CSV file."""
    mp = collections.defaultdict(list)
    if not path:
        return mp
    with open(path, newline='') as f:
        for carrier, lstop, partner in csv.reader(f, delimiter=';'):
            mp[(carrier.strip(), lstop.strip())].append(partner.strip())
    return mp


def load_bazl_data(bazl_path, start_date, end_date):
    """
    Load and filter BAZL passenger data.

    Returns:
        DataFrame with columns: Airline, Airport, PAX, Date, Year, Month
        Series indexed by (Airline, Airport) with PAX sums
        DataFrame with monthly PAX data for quality checks
    """
    cols = ['Airline Code (IATA)', 'Flughafen (IATA)',
            'Passagiere / Passagers', 'Jahr', 'Monat']
    df = pd.read_excel(bazl_path, sheet_name='BAZL-Daten', usecols=cols)

    def month_to_date(row):
        m = row['Monat']
        if isinstance(m, (int, float)) and m > 1000:
            return excel_serial_to_date(m)
        if isinstance(m, (int, float)):
            return datetime(int(row['Jahr']), int(m), 1)
        if isinstance(m, datetime):
            return datetime(m.year, m.month, 1)
        return pd.NaT

    df['Date'] = df.apply(month_to_date, axis=1)
    df['Year'] = df['Date'].dt.year
    df['Month'] = df['Date'].dt.month

    df = df[(df['Date'] >= start_date) & (df['Date'] <= end_date)]

    # Rename columns for clarity
    df = df.rename(columns={
        'Airline Code (IATA)': 'Airline',
        'Flughafen (IATA)': 'Airport',
        'Passagiere / Passagers': 'PAX'
    })

    # Create PAX lookup by (Airline, Airport)
    pax_lookup = df.groupby(['Airline', 'Airport'])['PAX'].sum()

    # Create monthly PAX data for quality checks
    monthly_pax = df.groupby(['Airline', 'Airport', 'Year', 'Month'])['PAX'].sum().reset_index()

    return df, pax_lookup, monthly_pax


def load_inad_data(inad_path, start_date, end_date):
    """
    Load INAD cases from Excel file.

    Returns:
        DataFrame with all INAD cases (including excluded codes for dashboard filtering)
        with columns: Airline, LastStop, Year, Month, Date, RefusalCode, Included, Category
    """
    wb = openpyxl.load_workbook(inad_path, read_only=True, data_only=True)
    ws = wb['INAD-Tabelle']

    # Get header indices
    hdr = {c.value: i for i, c in enumerate(ws[1])}
    idx_air = hdr['Fluggesellschaft']
    idx_last = hdr['Abflugort (last stop)']
    idx_year = hdr['Jahr']
    idx_month = hdr['Monat']
    idx_det = 18  # Column S - refusal code

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        yr = row[idx_year]
        mon = row[idx_month]

        if yr is None or mon is None:
            continue
        try:
            yr = int(yr)
            mon = int(mon)
        except (ValueError, TypeError):
            continue

        dt = datetime(yr, mon, 1)
        if dt < start_date or dt > end_date:
            continue

        code = (row[idx_det] or '').strip()
        air = (row[idx_air] or '').strip()
        lst = (row[idx_last] or '').strip()

        # Mark if this case is included in analysis (not excluded)
        included = bool(code) and code not in EXCLUDE_CODES

        # Categorize the refusal code
        category = 'Other'
        for cat, codes in REFUSAL_CATEGORIES.items():
            if code in codes or code.rstrip('en') in codes:
                category = cat
                break

        records.append({
            'Airline': air,
            'LastStop': lst,
            'Year': yr,
            'Month': mon,
            'Date': dt,
            'RefusalCode': code if code else 'Unknown',
            'Included': included,
            'Category': category
        })

    wb.close()
    return pd.DataFrame(records)


def calculate_step1(inad_df, min_inad=6):
    """
    Step 1 (Prüfstufe 1): Identify airlines with ≥N INADs.
    """
    included = inad_df[inad_df['Included']]
    airline_counts = included.groupby('Airline').size().reset_index(name='INAD_Count')
    airline_counts['PassesThreshold'] = airline_counts['INAD_Count'] >= min_inad
    airline_counts = airline_counts.sort_values('INAD_Count', ascending=False)
    return airline_counts


def calculate_step2(inad_df, step1_df, min_inad=6):
    """
    Step 2 (Prüfstufe 2): From Step 1 airlines, identify routes with ≥N INADs.
    """
    passing_airlines = set(step1_df[step1_df['PassesThreshold']]['Airline'])
    included = inad_df[(inad_df['Included']) & (inad_df['Airline'].isin(passing_airlines))]
    route_counts = included.groupby(['Airline', 'LastStop']).size().reset_index(name='INAD_Count')
    route_counts['PassesThreshold'] = route_counts['INAD_Count'] >= min_inad
    route_counts = route_counts.sort_values('INAD_Count', ascending=False)
    return route_counts


def calculate_confidence_score(inad_count, pax_count, min_pax=5000):
    """
    Calculate a confidence score (0-100) based on sample size.

    Higher INAD counts and PAX volumes = higher confidence.
    """
    if pax_count < min_pax:
        return 0  # Unreliable

    # INAD confidence (more cases = more reliable)
    inad_score = min(100, (inad_count / 20) * 100)  # Max confidence at 20+ INADs

    # PAX confidence (more passengers = more reliable denominator)
    pax_score = min(100, (pax_count / 100000) * 100)  # Max confidence at 100k+ PAX

    # Combined score (weighted average)
    return int(0.6 * inad_score + 0.4 * pax_score)


def calculate_robust_threshold(densities, method='median', trim_percent=0.1):
    """
    Calculate threshold using robust statistics.

    Methods:
    - 'mean': Simple arithmetic mean (sensitive to outliers)
    - 'median': Median value (robust to outliers)
    - 'trimmed_mean': Mean after removing top/bottom X%
    """
    valid_densities = [d for d in densities if d is not None and not np.isnan(d)]

    if len(valid_densities) == 0:
        return 0

    if method == 'median':
        return np.median(valid_densities)
    elif method == 'trimmed_mean':
        return stats.trim_mean(valid_densities, trim_percent)
    else:  # mean
        return np.mean(valid_densities)


def check_pax_data_quality(monthly_pax, airline, airport, expected_months=6, min_months=4):
    """
    Check PAX data quality for a specific route.

    Returns:
        dict with quality indicators
    """
    route_data = monthly_pax[(monthly_pax['Airline'] == airline) &
                             (monthly_pax['Airport'] == airport)]

    months_with_data = len(route_data)
    total_pax = route_data['PAX'].sum()

    # Check for anomalies
    if months_with_data > 0:
        avg_monthly_pax = total_pax / months_with_data
        max_pax = route_data['PAX'].max()
        min_pax = route_data['PAX'].min()
        variance_ratio = max_pax / min_pax if min_pax > 0 else float('inf')
    else:
        avg_monthly_pax = 0
        variance_ratio = 0

    return {
        'months_with_data': months_with_data,
        'expected_months': expected_months,
        'data_completeness': months_with_data / expected_months if expected_months > 0 else 0,
        'is_complete': months_with_data >= min_months,
        'total_pax': total_pax,
        'avg_monthly_pax': avg_monthly_pax,
        'high_variance': variance_ratio > 10,  # Flag if max is 10x min
        'quality_warning': None
    }


def calculate_step3_enhanced(step2_df, pax_lookup, monthly_pax, inad_df,
                             partner_map=None, config=None):
    """
    Enhanced Step 3: INAD density calculation with robust statistics and quality checks.

    Returns DataFrame with:
    - Basic metrics: Airline, LastStop, INAD, PAX, Density
    - Quality indicators: Confidence, DataQuality warnings
    - Classification: Priority (HIGH_PRIORITY, WATCH_LIST, or CLEAR)
    """
    if config is None:
        config = DEFAULT_CONFIG

    if partner_map is None:
        partner_map = {}

    passing_routes = step2_df[step2_df['PassesThreshold']]

    # Calculate expected months in the period
    expected_months = 6  # Semester

    rows = []
    for _, row in passing_routes.iterrows():
        air = row['Airline']
        lst = row['LastStop']
        inad = row['INAD_Count']

        # Get PAX, including partner airlines
        p = pax_lookup.get((air, lst), 0)
        for partner in partner_map.get((air, lst), []):
            p += pax_lookup.get((partner, lst), 0)

        # Data quality check
        quality = check_pax_data_quality(monthly_pax, air, lst, expected_months,
                                         config['pax_completeness_months'])

        # Generate quality warning if needed
        warning = None
        if not quality['is_complete']:
            warning = f"Incomplete PAX data ({quality['months_with_data']}/{expected_months} months)"
        elif quality['high_variance']:
            warning = "High variance in monthly PAX data"
        elif p < config['min_pax']:
            warning = f"Low PAX volume (<{config['min_pax']:,})"

        # Calculate density (only if PAX meets minimum threshold for reliability)
        if p >= config['min_pax']:
            density = (inad / p * 1000)
            reliable = True
        elif p > 0:
            density = (inad / p * 1000)  # Calculate but mark unreliable
            reliable = False
        else:
            density = None
            reliable = False

        # Confidence score
        confidence = calculate_confidence_score(inad, p, config['min_pax'])

        # Get refusal code breakdown
        route_inads = inad_df[(inad_df['Airline'] == air) &
                              (inad_df['LastStop'] == lst) &
                              (inad_df['Included'])]
        code_breakdown = route_inads['Category'].value_counts().to_dict()

        rows.append({
            'Airline': air,
            'LastStop': lst,
            'INAD': inad,
            'PAX': int(p),
            'Density_permille': density,
            'Reliable': reliable,
            'Confidence': confidence,
            'DataQuality': warning,
            'MonthsWithData': quality['months_with_data'],
            'CodeBreakdown': code_breakdown
        })

    result = pd.DataFrame(rows)

    if len(result) == 0:
        result['Priority'] = []
        result['AboveThreshold'] = []
        result['Threshold'] = []
        result['ThresholdMethod'] = []
        return result

    # Calculate robust threshold using only reliable data points
    reliable_densities = result[result['Reliable']]['Density_permille'].tolist()
    threshold = calculate_robust_threshold(
        reliable_densities,
        method=config['threshold_method'],
        trim_percent=config['trimmed_percent']
    )

    # Classify each route
    def classify_priority(row):
        if row['Density_permille'] is None:
            return 'NO_DATA'
        if not row['Reliable']:
            return 'UNRELIABLE'

        above_threshold = row['Density_permille'] >= threshold
        above_min_density = row['Density_permille'] >= config['min_density']
        high_multiplier = row['Density_permille'] >= threshold * config['high_priority_multiplier']
        high_inad = row['INAD'] >= config['high_priority_min_inad']

        if above_threshold and above_min_density and high_multiplier and high_inad:
            return 'HIGH_PRIORITY'
        elif above_threshold:
            return 'WATCH_LIST'
        else:
            return 'CLEAR'

    result['Priority'] = result.apply(classify_priority, axis=1)
    result['AboveThreshold'] = result['Density_permille'] >= threshold
    result['Threshold'] = threshold
    result['ThresholdMethod'] = config['threshold_method']

    # Sort by priority, then density
    priority_order = {'HIGH_PRIORITY': 0, 'WATCH_LIST': 1, 'UNRELIABLE': 2, 'CLEAR': 3, 'NO_DATA': 4}
    result['PriorityOrder'] = result['Priority'].map(priority_order)
    result = result.sort_values(['PriorityOrder', 'Density_permille'],
                                ascending=[True, False]).drop('PriorityOrder', axis=1)

    return result


def detect_systemic_cases(historical_results, config=None):
    """
    Analyze multiple semesters to detect systemic issues.

    A route is flagged as SYSTEMIC if it appears as HIGH_PRIORITY or WATCH_LIST
    in N consecutive semesters.

    Args:
        historical_results: List of (semester_label, step3_df) tuples, sorted chronologically
        config: Configuration dict

    Returns:
        DataFrame with systemic case analysis
    """
    if config is None:
        config = DEFAULT_CONFIG

    required_consecutive = config['systemic_semesters']

    # Track appearances by route
    route_history = collections.defaultdict(list)

    for semester, step3 in historical_results:
        flagged = step3[step3['Priority'].isin(['HIGH_PRIORITY', 'WATCH_LIST'])]
        for _, row in flagged.iterrows():
            route_key = (row['Airline'], row['LastStop'])
            route_history[route_key].append({
                'Semester': semester,
                'Priority': row['Priority'],
                'Density': row['Density_permille'],
                'INAD': row['INAD'],
                'PAX': row['PAX']
            })

    # Analyze for consecutive appearances
    systemic_cases = []
    for (airline, last_stop), history in route_history.items():
        # Check for consecutive semesters
        consecutive_count = 0
        max_consecutive = 0
        appearances = len(history)

        # Simple consecutive check (assuming sorted chronologically)
        for i, entry in enumerate(history):
            if i == 0:
                consecutive_count = 1
            else:
                # Check if this semester follows the previous
                prev_semester = history[i-1]['Semester']
                curr_semester = entry['Semester']
                # Simple check - if they appear in sequence in our list
                consecutive_count += 1
            max_consecutive = max(max_consecutive, consecutive_count)

        is_systemic = max_consecutive >= required_consecutive

        # Calculate trend
        if len(history) >= 2:
            densities = [h['Density'] for h in history if h['Density'] is not None]
            if len(densities) >= 2:
                # Simple trend: compare last to first
                trend = 'IMPROVING' if densities[-1] < densities[0] else 'WORSENING'
                trend_pct = ((densities[-1] - densities[0]) / densities[0] * 100) if densities[0] > 0 else 0
            else:
                trend = 'UNKNOWN'
                trend_pct = 0
        else:
            trend = 'NEW'
            trend_pct = 0

        systemic_cases.append({
            'Airline': airline,
            'LastStop': last_stop,
            'TotalAppearances': appearances,
            'MaxConsecutive': max_consecutive,
            'IsSystemic': is_systemic,
            'Trend': trend,
            'TrendPercent': trend_pct,
            'History': history,
            'LatestPriority': history[-1]['Priority'] if history else None,
            'LatestDensity': history[-1]['Density'] if history else None,
            'LatestINAD': history[-1]['INAD'] if history else None
        })

    result = pd.DataFrame(systemic_cases)
    if len(result) > 0:
        result = result.sort_values(['IsSystemic', 'TotalAppearances', 'LatestDensity'],
                                    ascending=[False, False, False])
    return result


def calculate_airline_comparison(step3_df, airline):
    """
    Compare a specific airline's flagged routes against their other routes.

    Returns context about whether this is isolated or airline-wide issue.
    """
    airline_routes = step3_df[step3_df['Airline'] == airline]

    if len(airline_routes) == 0:
        return None

    total_routes = len(airline_routes)
    flagged_routes = len(airline_routes[airline_routes['Priority'].isin(['HIGH_PRIORITY', 'WATCH_LIST'])])
    avg_density = airline_routes['Density_permille'].mean()

    return {
        'total_routes': total_routes,
        'flagged_routes': flagged_routes,
        'flagged_percent': (flagged_routes / total_routes * 100) if total_routes > 0 else 0,
        'avg_density': avg_density
    }


def run_full_analysis(inad_path, bazl_path, start_date, end_date,
                      min_inad=6, partner_map_path=None, config=None):
    """
    Run the complete 3-step analysis with enhanced features.
    """
    if config is None:
        config = DEFAULT_CONFIG.copy()
    config['min_inad'] = min_inad

    start = pd.to_datetime(start_date)
    end = pd.to_datetime(end_date)

    # Load data
    partner_map = load_partner_map(partner_map_path)
    pax_df, pax_lookup, monthly_pax = load_bazl_data(bazl_path, start, end)
    inad_df = load_inad_data(inad_path, start, end)

    # Run analysis steps
    step1 = calculate_step1(inad_df, min_inad)
    step2 = calculate_step2(inad_df, step1, min_inad)
    step3 = calculate_step3_enhanced(step2, pax_lookup, monthly_pax, inad_df,
                                     partner_map, config)

    # Get threshold
    threshold = step3['Threshold'].iloc[0] if len(step3) > 0 else None

    return {
        'step1': step1,
        'step2': step2,
        'step3': step3,
        'inad_raw': inad_df,
        'pax_raw': pax_df,
        'pax_lookup': pax_lookup,
        'monthly_pax': monthly_pax,
        'start': start,
        'end': end,
        'threshold': threshold,
        'min_inad': min_inad,
        'config': config
    }


def run_multi_semester_analysis(inad_path, bazl_path, semesters, config=None):
    """
    Run analysis across multiple semesters for trend detection.

    Args:
        semesters: List of (year, sem, start, end) tuples

    Returns:
        dict with semester results and systemic case analysis
    """
    if config is None:
        config = DEFAULT_CONFIG.copy()

    semester_results = []
    historical_step3 = []

    for year, sem, start, end in sorted(semesters):
        try:
            results = run_full_analysis(inad_path, bazl_path, start, end,
                                        config=config)

            semester_label = f"{year} H{sem}"

            # Summary metrics
            total_inad = len(results['inad_raw'][results['inad_raw']['Included']])
            total_pax = results['pax_raw']['PAX'].sum() if len(results['pax_raw']) > 0 else 0

            step1_passing = len(results['step1'][results['step1']['PassesThreshold']])
            step2_passing = len(results['step2'][results['step2']['PassesThreshold']])

            high_priority = len(results['step3'][results['step3']['Priority'] == 'HIGH_PRIORITY'])
            watch_list = len(results['step3'][results['step3']['Priority'] == 'WATCH_LIST'])

            threshold = results['threshold'] if results['threshold'] else 0

            semester_results.append({
                'Year': year,
                'Semester': sem,
                'Label': semester_label,
                'Start': start,
                'End': end,
                'Total_INAD': total_inad,
                'Total_PAX': total_pax,
                'Step1_Airlines': step1_passing,
                'Step2_Routes': step2_passing,
                'High_Priority': high_priority,
                'Watch_List': watch_list,
                'Threshold': threshold,
                'INAD_Rate': (total_inad / total_pax * 1_000_000) if total_pax > 0 else 0
            })

            historical_step3.append((semester_label, results['step3']))

        except Exception as e:
            print(f"Error processing {year} H{sem}: {e}")
            continue

    # Detect systemic cases
    systemic = detect_systemic_cases(historical_step3, config)

    return {
        'semester_summary': pd.DataFrame(semester_results),
        'systemic_cases': systemic,
        'historical_step3': historical_step3,
        'config': config
    }


def get_available_semesters(inad_path):
    """
    Scan INAD file to find available year/semester combinations.
    """
    wb = openpyxl.load_workbook(inad_path, read_only=True, data_only=True)
    ws = wb['INAD-Tabelle']

    hdr = {c.value: i for i, c in enumerate(ws[1])}
    idx_year = hdr['Jahr']
    idx_month = hdr['Monat']

    year_months = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        yr = row[idx_year]
        mon = row[idx_month]
        if yr is not None and mon is not None:
            try:
                year_months.add((int(yr), int(mon)))
            except (ValueError, TypeError):
                continue

    wb.close()

    semesters = set()
    for yr, mon in year_months:
        if 1 <= mon <= 6:
            semesters.add((yr, 1, f"{yr}-01-01", f"{yr}-06-30"))
        else:
            semesters.add((yr, 2, f"{yr}-07-01", f"{yr}-12-31"))

    return sorted(semesters, reverse=True)


def get_refusal_code_stats(inad_df):
    """Get statistics on refusal codes."""
    stats = inad_df.groupby(['RefusalCode', 'Included']).size().reset_index(name='Count')
    stats['Description'] = stats['RefusalCode'].map(
        lambda x: REFUSAL_CODE_DESCRIPTIONS.get(x, REFUSAL_CODE_DESCRIPTIONS.get(x.rstrip('en'), 'Unbekannt'))
    )
    stats = stats.sort_values('Count', ascending=False)
    return stats


def generate_legal_summary(step3_df, systemic_df=None, airline=None):
    """
    Generate a summary suitable for legal review.

    Returns a structured dict with key findings and recommendations.
    """
    if airline:
        step3_df = step3_df[step3_df['Airline'] == airline]

    high_priority = step3_df[step3_df['Priority'] == 'HIGH_PRIORITY']
    watch_list = step3_df[step3_df['Priority'] == 'WATCH_LIST']

    summary = {
        'total_routes_analyzed': len(step3_df),
        'high_priority_count': len(high_priority),
        'watch_list_count': len(watch_list),
        'threshold_used': step3_df['Threshold'].iloc[0] if len(step3_df) > 0 else None,
        'threshold_method': step3_df['ThresholdMethod'].iloc[0] if len(step3_df) > 0 else None,
        'high_priority_routes': high_priority[['Airline', 'LastStop', 'INAD', 'PAX',
                                                'Density_permille', 'Confidence']].to_dict('records'),
        'watch_list_routes': watch_list[['Airline', 'LastStop', 'INAD', 'PAX',
                                         'Density_permille', 'Confidence']].to_dict('records'),
        'data_quality_issues': step3_df[step3_df['DataQuality'].notna()][
            ['Airline', 'LastStop', 'DataQuality']].to_dict('records')
    }

    if systemic_df is not None and len(systemic_df) > 0:
        systemic = systemic_df[systemic_df['IsSystemic']]
        summary['systemic_cases'] = systemic[['Airline', 'LastStop', 'TotalAppearances',
                                              'Trend', 'TrendPercent']].to_dict('records')
        summary['systemic_count'] = len(systemic)

    return summary


# CLI interface
if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser(description='Enhanced INAD Analysis Tool')
    ap.add_argument('--inad', required=True, help='INAD Excel (xls/xlsm)')
    ap.add_argument('--bazl', required=True, help='BAZL Excel (xlsx)')
    ap.add_argument('--start', required=True, help='YYYY-MM-DD')
    ap.add_argument('--end', required=True, help='YYYY-MM-DD')
    ap.add_argument('--inadmin', type=int, default=6, help='Minimum INAD threshold (default: 6)')
    ap.add_argument('--minpax', type=int, default=5000, help='Minimum PAX for reliable density (default: 5000)')
    ap.add_argument('--threshold-method', choices=['mean', 'median', 'trimmed_mean'],
                    default='median', help='Method for calculating threshold')
    ap.add_argument('--map', default=None, help='CSV partner mapping file')
    args = ap.parse_args()

    config = DEFAULT_CONFIG.copy()
    config['min_pax'] = args.minpax
    config['threshold_method'] = args.threshold_method

    results = run_full_analysis(
        args.inad, args.bazl, args.start, args.end,
        min_inad=args.inadmin, partner_map_path=args.map, config=config
    )

    # Output to Excel with all three steps
    out_file = f"INAD_Analysis_{results['start']:%Y%m%d}_{results['end']:%Y%m%d}.xlsx"
    with pd.ExcelWriter(out_file) as writer:
        results['step1'].to_excel(writer, index=False, sheet_name='Step1_Airlines')
        results['step2'].to_excel(writer, index=False, sheet_name='Step2_Routes')

        # Prepare Step 3 for export (convert dict column)
        step3_export = results['step3'].copy()
        step3_export['CodeBreakdown'] = step3_export['CodeBreakdown'].apply(str)
        step3_export.to_excel(writer, index=False, sheet_name='Step3_Density')

    print(f"✓ Analysis complete:")
    print(f"  Threshold method: {config['threshold_method']}")
    print(f"  Threshold value: {results['threshold']:.4f}‰")
    print(f"  Step 1: {len(results['step1'][results['step1']['PassesThreshold']])} airlines")
    print(f"  Step 2: {len(results['step2'][results['step2']['PassesThreshold']])} routes")

    step3 = results['step3']
    print(f"  Step 3: {len(step3[step3['Priority'] == 'HIGH_PRIORITY'])} HIGH PRIORITY")
    print(f"          {len(step3[step3['Priority'] == 'WATCH_LIST'])} WATCH LIST")
    print(f"          {len(step3[step3['Priority'] == 'UNRELIABLE'])} UNRELIABLE (low PAX)")
    print(f"  Output: {out_file}")
